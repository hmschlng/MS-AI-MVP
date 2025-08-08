"""
Test module for excel_generator.py

Tests for Excel workbook generation functionality
"""
import pytest
import pandas as pd
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from unittest.mock import Mock, patch

from ai_test_generator.excel.excel_generator import ExcelGenerator
from ai_test_generator.excel.excel_templates import (
    ExcelTestScenario, TestPriority, TestType, TestStatus
)
from ai_test_generator.excel.excel_validator import ValidationResult, ValidationError
from ai_test_generator.core.llm_agent import TestScenario


class TestExcelGenerator:
    """Test ExcelGenerator class"""
    
    @pytest.fixture
    def generator(self):
        """ExcelGenerator instance fixture"""
        return ExcelGenerator()
    
    @pytest.fixture
    def sample_excel_scenarios(self):
        """Sample ExcelTestScenario list fixture"""
        return [
            ExcelTestScenario(
                scenario_id="TC001",
                feature="User Authentication",
                description="Test login functionality",
                preconditions="User has valid account",
                test_steps="1. Navigate to login\n2. Enter credentials\n3. Click login",
                expected_results="User is logged in successfully",
                test_data="username: test@example.com",
                priority=TestPriority.HIGH.value,
                test_type=TestType.FUNCTIONAL.value,
                status=TestStatus.NOT_EXECUTED.value
            ),
            ExcelTestScenario(
                scenario_id="TC002",
                feature="User Profile",
                description="Test profile update",
                preconditions="User is logged in",
                test_steps="1. Go to profile\n2. Update info\n3. Save changes",
                expected_results="Profile is updated",
                priority=TestPriority.MEDIUM.value,
                test_type=TestType.FUNCTIONAL.value,
                status=TestStatus.NOT_EXECUTED.value
            )
        ]
    
    @pytest.fixture
    def sample_llm_scenarios(self):
        """Sample TestScenario list fixture"""
        return [
            TestScenario(
                scenario_id="TS001",
                feature="Data Validation",
                description="Test input validation",
                preconditions=["System is ready", "Valid data exists"],
                test_steps=[
                    {"action": "Enter valid data", "data": "test input"},
                    {"action": "Submit form", "data": ""}
                ],
                expected_results=["Data is validated", "Form is submitted"],
                test_data={"input": "test data"},
                priority="High",
                test_type="Functional"
            )
        ]
    
    def test_generator_initialization(self, generator):
        """Test ExcelGenerator initialization"""
        assert generator.validator is not None
        assert generator.styles is not None
        assert generator.template is not None
    
    def test_generate_from_llm_scenarios(self, generator, sample_llm_scenarios):
        """Test generating workbook from LLM scenarios"""
        wb = generator.generate_from_llm_scenarios(sample_llm_scenarios)
        
        # Check workbook is created
        assert isinstance(wb, Workbook)
        
        # Check expected sheets exist
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "Test Scenarios" in sheet_names
        assert "Template" in sheet_names
        assert "Validation" in sheet_names
        
        # Check Test Scenarios sheet has data
        scenarios_sheet = wb["Test Scenarios"]
        assert scenarios_sheet is not None
        
        # Check first row is header
        headers = [cell.value for cell in scenarios_sheet[1]]
        assert "Scenario ID" in headers
        assert "Feature" in headers
        assert "Description" in headers
        
        # Check data row exists
        assert scenarios_sheet['A2'].value == "TC001"  # Generated ID
    
    def test_generate_workbook_with_project_info(self, generator, sample_excel_scenarios):
        """Test workbook generation with project info"""
        project_info = {
            "project_name": "Test Project",
            "version": "2.0.0",
            "test_environment": "Staging",
            "tester": "Test Engineer"
        }
        
        wb = generator.generate_workbook(sample_excel_scenarios, project_info)
        
        # Check Summary sheet has project info
        summary_sheet = wb["Summary"]
        
        # Find project name cell (it should be in the summary)
        found_project_name = False
        for row in summary_sheet.iter_rows():
            for cell in row:
                if cell.value == "Test Project":
                    found_project_name = True
                    break
            if found_project_name:
                break
        
        assert found_project_name, "Project name not found in summary sheet"
    
    def test_create_scenarios_sheet(self, generator, sample_excel_scenarios):
        """Test scenarios sheet creation"""
        wb = Workbook()
        ws = wb.active
        
        generator._create_scenarios_sheet(ws, sample_excel_scenarios)
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            "Scenario ID", "Feature", "Description", "Preconditions",
            "Test Steps", "Expected Results", "Test Data", "Priority",
            "Test Type", "Status", "Assigned To", "Estimated Time (min)",
            "Actual Time (min)", "Notes"
        ]
        
        for header in expected_headers:
            assert header in headers
        
        # Check data
        assert ws['A2'].value == "TC001"
        assert ws['B2'].value == "User Authentication"
        assert ws['A3'].value == "TC002"
        assert ws['B3'].value == "User Profile"
        
        # Check styling is applied (header should have fill)
        assert ws['A1'].fill is not None
        assert ws['A1'].font is not None
    
    def test_create_summary_sheet(self, generator, sample_excel_scenarios):
        """Test summary sheet creation"""
        wb = Workbook()
        ws = wb.active
        
        generator._create_summary_sheet(ws, sample_excel_scenarios, None)
        
        # Check title exists
        assert ws['A1'].value == "Test Execution Summary"
        
        # Check sections exist by looking for section titles
        found_sections = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in ["Project Information", "Test Statistics", "Priority Breakdown"]:
                    found_sections.append(cell.value)
        
        assert "Project Information" in found_sections
        assert "Test Statistics" in found_sections
        assert "Priority Breakdown" in found_sections
    
    def test_create_template_sheet(self, generator):
        """Test template sheet creation"""
        wb = Workbook()
        ws = wb.active
        
        generator._create_template_sheet(ws)
        
        # Check headers exist
        headers = [cell.value for cell in ws[1]]
        assert "Scenario ID" in headers
        assert "Feature" in headers
        
        # Check default values in some cells
        assert ws['H2'].value == TestPriority.MEDIUM.value  # Priority column
        assert ws['I2'].value == TestType.FUNCTIONAL.value  # Test Type column
        assert ws['J2'].value == TestStatus.NOT_EXECUTED.value  # Status column
    
    @patch('ai_test_generator.excel.excel_generator.ExcelValidator')
    def test_create_validation_sheet(self, mock_validator_class, generator, sample_excel_scenarios):
        """Test validation sheet creation"""
        # Mock validation result
        mock_result = ValidationResult(
            is_valid=False,
            errors=[
                ValidationError(
                    row_index=0,
                    column="Scenario ID",
                    error_type="required",
                    message="Scenario ID is required",
                    severity="error"
                )
            ],
            warnings=[
                ValidationError(
                    row_index=1,
                    column="Description",
                    error_type="length",
                    message="Description is too long",
                    severity="warning"
                )
            ],
            total_scenarios=2,
            valid_scenarios=1
        )
        
        mock_validator = Mock()
        mock_validator.validate_scenarios.return_value = mock_result
        generator.validator = mock_validator
        
        wb = Workbook()
        ws = wb.active
        
        generator._create_validation_sheet(ws, sample_excel_scenarios)
        
        # Check title
        assert ws['A1'].value == "Validation Results"
        
        # Check summary info exists
        found_summary = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Summary":
                    found_summary = True
                    break
            if found_summary:
                break
        
        assert found_summary
        
        # Check that validation was called
        mock_validator.validate_scenarios.assert_called_once_with(sample_excel_scenarios)
    
    def test_adjust_column_widths(self, generator):
        """Test column width adjustment"""
        wb = Workbook()
        ws = wb.active
        
        # Add some headers
        headers = ["Scenario ID", "Feature", "Description"]
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)
        
        generator._adjust_column_widths(ws)
        
        # Check that column widths are set
        assert ws.column_dimensions['A'].width > 0
        assert ws.column_dimensions['B'].width > 0
        assert ws.column_dimensions['C'].width > 0
    
    def test_add_data_validations(self, generator):
        """Test data validation rules"""
        wb = Workbook()
        ws = wb.active
        
        generator._add_data_validations(ws, 5)
        
        # Check that data validations are added
        assert len(ws.data_validations.dataValidation) > 0
    
    def test_save_workbook(self, generator, sample_excel_scenarios):
        """Test saving workbook to file"""
        wb = generator.generate_workbook(sample_excel_scenarios)
        
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "test_workbook.xlsx"
            
            result_path = generator.save_workbook(wb, file_path)
            
            # Check file was created
            assert Path(result_path).exists()
            assert result_path == str(file_path)
    
    def test_export_to_bytes(self, generator, sample_excel_scenarios):
        """Test exporting workbook to bytes"""
        wb = generator.generate_workbook(sample_excel_scenarios)
        
        byte_data = generator.export_to_bytes(wb)
        
        # Check we got bytes back
        assert isinstance(byte_data, bytes)
        assert len(byte_data) > 0
    
    def test_create_editable_dataframe(self, generator, sample_excel_scenarios):
        """Test creating editable DataFrame"""
        df = generator.create_editable_dataframe(sample_excel_scenarios)
        
        # Check DataFrame structure
        assert isinstance(df, pd.DataFrame)
        assert len(df) == len(sample_excel_scenarios)
        
        # Check columns exist
        expected_columns = [
            "Scenario ID", "Feature", "Description", "Priority"
        ]
        for col in expected_columns:
            assert col in df.columns
        
        # Check data
        assert df.iloc[0]["Scenario ID"] == "TC001"
        assert df.iloc[1]["Scenario ID"] == "TC002"
    
    def test_create_editable_dataframe_empty(self, generator):
        """Test creating editable DataFrame with empty scenarios"""
        df = generator.create_editable_dataframe([])
        
        # Should return template dataframe
        assert isinstance(df, pd.DataFrame)
        assert len(df) == 1  # Template has one sample row
        assert df.iloc[0]["Scenario ID"] == "TC001"  # Sample data
    
    def test_dataframe_to_scenarios(self, generator):
        """Test converting DataFrame to scenarios"""
        data = {
            "Scenario ID": ["TC001", "TC002"],
            "Feature": ["Login", "Profile"],
            "Description": ["Test login", "Test profile"],
            "Preconditions": ["", "User logged in"],
            "Test Steps": ["1. Login", "1. Update profile"],
            "Expected Results": ["Success", "Updated"],
            "Test Data": ["", ""],
            "Priority": [TestPriority.HIGH.value, TestPriority.MEDIUM.value],
            "Test Type": [TestType.FUNCTIONAL.value, TestType.FUNCTIONAL.value],
            "Status": [TestStatus.NOT_EXECUTED.value, TestStatus.NOT_EXECUTED.value],
            "Assigned To": ["", ""],
            "Estimated Time (min)": ["", ""],
            "Actual Time (min)": ["", ""],
            "Notes": ["", ""]
        }
        
        df = pd.DataFrame(data)
        scenarios = generator.dataframe_to_scenarios(df)
        
        assert len(scenarios) == 2
        assert scenarios[0].scenario_id == "TC001"
        assert scenarios[1].scenario_id == "TC002"
        assert scenarios[0].priority == TestPriority.HIGH.value
    
    def test_generate_from_dataframe(self, generator):
        """Test generating workbook from DataFrame"""
        data = {
            "Scenario ID": ["TC001"],
            "Feature": ["Test Feature"],
            "Description": ["Test Description"],
            "Preconditions": [""],
            "Test Steps": ["1. Test step"],
            "Expected Results": ["Success"],
            "Test Data": [""],
            "Priority": [TestPriority.MEDIUM.value],
            "Test Type": [TestType.FUNCTIONAL.value],
            "Status": [TestStatus.NOT_EXECUTED.value],
            "Assigned To": [""],
            "Estimated Time (min)": [""],
            "Actual Time (min)": [""],
            "Notes": [""]
        }
        
        df = pd.DataFrame(data)
        wb = generator.generate_from_dataframe(df)
        
        # Check workbook is created
        assert isinstance(wb, Workbook)
        
        # Check data is in scenarios sheet
        scenarios_sheet = wb["Test Scenarios"]
        assert scenarios_sheet['A2'].value == "TC001"
        assert scenarios_sheet['B2'].value == "Test Feature"
    
    def test_get_default_project_info(self, generator):
        """Test getting default project info"""
        project_info = generator.get_default_project_info()
        
        assert isinstance(project_info, dict)
        assert "project_name" in project_info
        assert "version" in project_info
        assert "test_environment" in project_info
        assert "test_period" in project_info
        assert "tester" in project_info
        
        assert project_info["project_name"] == "AI Generated Test Project"
        assert project_info["version"] == "1.0.0"


class TestExcelGeneratorIntegration:
    """Integration tests for ExcelGenerator"""
    
    def test_full_workflow_llm_to_excel(self):
        """Test complete workflow from LLM scenarios to Excel file"""
        generator = ExcelGenerator()
        
        # Create LLM scenarios
        llm_scenarios = [
            TestScenario(
                scenario_id="LLM001",
                feature="API Testing",
                description="Test REST API endpoints",
                preconditions=["API server running", "Valid authentication"],
                test_steps=[
                    {"action": "Send GET request", "data": "/api/users"},
                    {"action": "Verify response", "data": "status 200"}
                ],
                expected_results=["Valid JSON response", "User list returned"],
                test_data={"endpoint": "/api/users"},
                priority="High",
                test_type="Integration"
            )
        ]
        
        project_info = {
            "project_name": "API Test Suite",
            "version": "1.0.0",
            "test_environment": "Development"
        }
        
        # Generate workbook
        wb = generator.generate_from_llm_scenarios(llm_scenarios, project_info)
        
        # Verify complete workbook
        assert isinstance(wb, Workbook)
        assert len(wb.sheetnames) == 4
        
        # Check each sheet
        summary_sheet = wb["Summary"]
        scenarios_sheet = wb["Test Scenarios"]
        template_sheet = wb["Template"]
        validation_sheet = wb["Validation"]
        
        assert summary_sheet is not None
        assert scenarios_sheet is not None
        assert template_sheet is not None
        assert validation_sheet is not None
        
        # Check scenarios data
        assert scenarios_sheet['A2'].value == "TC001"  # Auto-generated ID
        assert scenarios_sheet['B2'].value == "API Testing"
        
        # Save and verify file can be created
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "integration_test.xlsx"
            saved_path = generator.save_workbook(wb, file_path)
            assert Path(saved_path).exists()
    
    def test_excel_roundtrip(self):
        """Test Excel -> DataFrame -> Excel roundtrip"""
        generator = ExcelGenerator()
        
        original_scenarios = [
            ExcelTestScenario(
                scenario_id="RT001",
                feature="Roundtrip Test",
                description="Test data integrity",
                preconditions="Test setup",
                test_steps="1. Process data",
                expected_results="Data intact",
                priority=TestPriority.LOW.value,
                test_type=TestType.REGRESSION.value
            )
        ]
        
        # Create DataFrame
        df = generator.create_editable_dataframe(original_scenarios)
        
        # Convert back to scenarios
        converted_scenarios = generator.dataframe_to_scenarios(df)
        
        # Verify data integrity
        assert len(converted_scenarios) == len(original_scenarios)
        original = original_scenarios[0]
        converted = converted_scenarios[0]
        
        assert original.scenario_id == converted.scenario_id
        assert original.feature == converted.feature
        assert original.description == converted.description
        assert original.priority == converted.priority
        assert original.test_type == converted.test_type
    
    @patch('ai_test_generator.excel.excel_generator.logger')
    def test_error_handling_in_dataframe_conversion(self, mock_logger):
        """Test error handling during DataFrame to scenarios conversion"""
        generator = ExcelGenerator()
        
        # Create DataFrame with invalid data - using objects that can't be stringified
        class BadObject:
            def __str__(self):
                raise ValueError("Cannot convert to string")
        
        invalid_data = {
            "Scenario ID": [BadObject(), "TC002"],  # Object that causes error on str()
            "Feature": ["Valid Feature", "Another Feature"],
        }
        
        df = pd.DataFrame(invalid_data)
        
        # This should handle errors gracefully
        scenarios = generator.dataframe_to_scenarios(df)
        
        # Should get fewer scenarios than rows due to conversion errors
        assert len(scenarios) <= len(df)
        
        # Should have logged warnings for the failed conversion
        assert mock_logger.warning.called