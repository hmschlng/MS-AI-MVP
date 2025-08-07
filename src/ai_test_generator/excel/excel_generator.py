"""
Excel Generator Module - 엑셀 문서 생성 핵심 로직

테스트 시나리오를 엑셀 문서로 생성하고, 
Streamlit에서 편집 가능한 형태로 변환합니다.
"""
import os
import io
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .excel_templates import (
    ExcelTestScenario, ExcelStyles, ExcelTemplate, 
    TestPriority, TestType, TestStatus
)
from .excel_validator import ExcelValidator, ValidationResult
from ai_test_generator.core.llm_agent import TestScenario
from ai_test_generator.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelGenerator:
    """엑셀 문서 생성 및 관리 클래스"""
    
    def __init__(self):
        self.validator = ExcelValidator()
        self.styles = ExcelStyles()
        self.template = ExcelTemplate()
    
    def generate_from_llm_scenarios(
        self, 
        scenarios: List[TestScenario], 
        project_info: Optional[Dict[str, str]] = None
    ) -> Workbook:
        """LLM 생성 시나리오를 엑셀로 변환"""
        logger.info(f"Converting {len(scenarios)} LLM scenarios to Excel format")
        
        # TestScenario를 ExcelTestScenario로 변환
        excel_scenarios = []
        for i, scenario in enumerate(scenarios):
            scenario_id = f"TC{i+1:03d}"  # TC001, TC002, ...
            excel_scenario = ExcelTestScenario.from_test_scenario(scenario, scenario_id)
            excel_scenarios.append(excel_scenario)
        
        return self.generate_workbook(excel_scenarios, project_info)
    
    def generate_workbook(
        self, 
        scenarios: List[ExcelTestScenario], 
        project_info: Optional[Dict[str, str]] = None
    ) -> Workbook:
        """엑셀 워크북 생성 (멀티 시트)"""
        logger.info(f"Generating Excel workbook with {len(scenarios)} scenarios")
        
        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거
        
        # 1. 테스트 시나리오 시트
        scenarios_sheet = wb.create_sheet("Test Scenarios")
        self._create_scenarios_sheet(scenarios_sheet, scenarios)
        
        # 2. 요약 시트
        summary_sheet = wb.create_sheet("Summary", 0)  # 첫 번째 시트로 설정
        self._create_summary_sheet(summary_sheet, scenarios, project_info)
        
        # 3. 템플릿 시트 (빈 템플릿)
        template_sheet = wb.create_sheet("Template")
        self._create_template_sheet(template_sheet)
        
        # 4. 검증 결과 시트
        validation_sheet = wb.create_sheet("Validation")
        self._create_validation_sheet(validation_sheet, scenarios)
        
        logger.info("Excel workbook generated successfully")
        return wb
    
    def _create_scenarios_sheet(self, ws: Worksheet, scenarios: List[ExcelTestScenario]):
        """테스트 시나리오 시트 생성"""
        # DataFrame 생성
        data = [scenario.to_dict() for scenario in scenarios]
        df = pd.DataFrame(data)
        
        # 데이터를 시트에 추가
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 헤더 스타일 적용
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.center_alignment
            cell.border = self.styles.thin_border
        
        # 컬럼 너비 자동 조정
        self._adjust_column_widths(ws)
        
        # 데이터 영역 스타일 적용
        for row_num in range(2, len(scenarios) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.wrap_alignment
                
                # Priority 기반 색상 적용
                if df.columns[col_num - 1] == "Priority":
                    priority = cell.value
                    if priority == TestPriority.HIGH.value:
                        cell.fill = self.styles.priority_high_fill
                    elif priority == TestPriority.MEDIUM.value:
                        cell.fill = self.styles.priority_medium_fill
                    elif priority == TestPriority.LOW.value:
                        cell.fill = self.styles.priority_low_fill
                
                # Status 기반 색상 적용
                elif df.columns[col_num - 1] == "Status":
                    status = cell.value
                    if status == TestStatus.PASS.value:
                        cell.fill = self.styles.status_pass_fill
                    elif status == TestStatus.FAIL.value:
                        cell.fill = self.styles.status_fail_fill
                    elif status == TestStatus.BLOCKED.value:
                        cell.fill = self.styles.status_blocked_fill
        
        # 데이터 검증 추가
        self._add_data_validations(ws, len(scenarios))
        
        # 시트 보호 설정 (편집 가능한 셀만 열어둠)
        ws.protection.sheet = True
        ws.protection.password = "testscenarios"
        
        # 편집 가능한 셀들의 보호 해제
        for row_num in range(2, len(scenarios) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.protection = Protection(locked=False)
    
    def _create_summary_sheet(self, ws: Worksheet, scenarios: List[ExcelTestScenario], project_info: Optional[Dict[str, str]]):
        """요약 시트 생성"""
        summary_template = self.template.get_summary_template()
        
        # 제목
        ws['A1'] = summary_template['title']
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws['A1'].alignment = self.styles.center_alignment
        ws.merge_cells('A1:D1')
        
        current_row = 3
        
        # 프로젝트 정보 섹션
        for section in summary_template['sections']:
            # 섹션 제목
            ws[f'A{current_row}'] = section['name']
            ws[f'A{current_row}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'A{current_row}'].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
            
            # 섹션 필드들
            for field_name, field_value in section['fields']:
                ws[f'A{current_row}'] = field_name
                ws[f'A{current_row}'].font = Font(name='Arial', size=10, bold=True)
                
                # 프로젝트 정보가 있으면 사용, 아니면 기본값/수식 사용
                if section['name'] == "Project Information" and project_info:
                    value = project_info.get(field_name.lower().replace(' ', '_'), field_value)
                else:
                    value = field_value
                
                ws[f'B{current_row}'] = value
                ws[f'B{current_row}'].font = Font(name='Arial', size=10)
                current_row += 1
            
            current_row += 1  # 섹션 간 공백
        
        # 차트 영역 추가
        self._add_summary_charts(ws, current_row)
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_template_sheet(self, ws: Worksheet):
        """빈 템플릿 시트 생성"""
        # 컬럼 헤더만 추가
        columns = [col["field"] for col in self.template.get_column_definitions()]
        
        for col_num, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.center_alignment
            cell.border = self.styles.thin_border
        
        # 컬럼 너비 조정
        self._adjust_column_widths(ws)
        
        # 빈 행 5개 추가 (예시용)
        for row_num in range(2, 7):
            for col_num in range(1, len(columns) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = self.styles.thin_border
                if col_num == 8:  # Priority 컬럼
                    cell.value = TestPriority.MEDIUM.value
                elif col_num == 9:  # Test Type 컬럼
                    cell.value = TestType.FUNCTIONAL.value
                elif col_num == 10:  # Status 컬럼
                    cell.value = TestStatus.NOT_EXECUTED.value
        
        # 데이터 검증 추가
        self._add_data_validations(ws, 100)  # 100행까지 검증 적용
    
    def _create_validation_sheet(self, ws: Worksheet, scenarios: List[ExcelTestScenario]):
        """검증 결과 시트 생성"""
        # 검증 실행
        validation_result = self.validator.validate_scenarios(scenarios)
        
        # 제목
        ws['A1'] = "Validation Results"
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # 요약 정보
        ws['A3'] = "Summary"
        ws['A3'].font = Font(name='Arial', size=12, bold=True)
        
        ws['A4'] = "Total Scenarios:"
        ws['B4'] = validation_result.total_scenarios
        ws['A5'] = "Valid Scenarios:"
        ws['B5'] = validation_result.valid_scenarios
        ws['A6'] = "Errors:"
        ws['B6'] = validation_result.error_count
        ws['A7'] = "Warnings:"
        ws['B7'] = validation_result.warning_count
        
        # 오류 목록
        if validation_result.errors or validation_result.warnings:
            ws['A9'] = "Issues Found"
            ws['A9'].font = Font(name='Arial', size=12, bold=True)
            
            # 헤더
            headers = ["Row", "Column", "Type", "Severity", "Message"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=10, column=col_num, value=header)
                cell.font = self.styles.header_font
                cell.fill = self.styles.header_fill
            
            # 오류/경고 데이터
            current_row = 11
            for error in validation_result.errors + validation_result.warnings:
                ws.cell(row=current_row, column=1, value=error.row_index + 1)
                ws.cell(row=current_row, column=2, value=error.column)
                ws.cell(row=current_row, column=3, value=error.error_type)
                ws.cell(row=current_row, column=4, value=error.severity)
                ws.cell(row=current_row, column=5, value=error.message)
                
                # 심각도에 따른 색상 적용
                for col in range(1, 6):
                    cell = ws.cell(row=current_row, column=col)
                    if error.severity == "error":
                        cell.fill = self.styles.status_fail_fill
                    elif error.severity == "warning":
                        cell.fill = self.styles.status_blocked_fill
                
                current_row += 1
        
        # 컬럼 너비 조정
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
    
    def _adjust_column_widths(self, ws: Worksheet):
        """컬럼 너비 자동 조정"""
        column_widths = {
            'A': 15,  # Scenario ID
            'B': 20,  # Feature
            'C': 30,  # Description
            'D': 25,  # Preconditions
            'E': 35,  # Test Steps
            'F': 25,  # Expected Results
            'G': 20,  # Test Data
            'H': 12,  # Priority
            'I': 15,  # Test Type
            'J': 12,  # Status
            'K': 15,  # Assigned To
            'L': 18,  # Estimated Time
            'M': 15,  # Actual Time
            'N': 25   # Notes
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _add_data_validations(self, ws: Worksheet, max_row: int):
        """데이터 검증 규칙 추가"""
        # Priority 검증
        priority_validation = DataValidation(
            type="list",
            formula1=f'"{",".join([e.value for e in TestPriority])}"',
            allow_blank=True
        )
        priority_validation.error = "올바른 우선순위를 선택하세요"
        priority_validation.errorTitle = "잘못된 우선순위"
        ws.add_data_validation(priority_validation)
        priority_validation.add(f'H2:H{max_row + 1}')
        
        # Test Type 검증
        type_validation = DataValidation(
            type="list",
            formula1=f'"{",".join([e.value for e in TestType])}"',
            allow_blank=True
        )
        type_validation.error = "올바른 테스트 타입을 선택하세요"
        type_validation.errorTitle = "잘못된 테스트 타입"
        ws.add_data_validation(type_validation)
        type_validation.add(f'I2:I{max_row + 1}')
        
        # Status 검증
        status_validation = DataValidation(
            type="list",
            formula1=f'"{",".join([e.value for e in TestStatus])}"',
            allow_blank=True
        )
        status_validation.error = "올바른 상태를 선택하세요"
        status_validation.errorTitle = "잘못된 상태"
        ws.add_data_validation(status_validation)
        status_validation.add(f'J2:J{max_row + 1}')
    
    def _add_summary_charts(self, ws: Worksheet, start_row: int):
        """요약 차트 추가 (데이터만, 실제 차트는 수동 추가 필요)"""
        ws[f'A{start_row}'] = "Chart Data"
        ws[f'A{start_row}'].font = Font(name='Arial', size=12, bold=True)
        
        # 상태별 통계 차트 데이터
        chart_data = [
            ["Status", "Count"],
            ["Pass", "=COUNTIF('Test Scenarios'!J:J,\"Pass\")"],
            ["Fail", "=COUNTIF('Test Scenarios'!J:J,\"Fail\")"],
            ["Blocked", "=COUNTIF('Test Scenarios'!J:J,\"Blocked\")"],
            ["Not Executed", "=COUNTIF('Test Scenarios'!J:J,\"Not Executed\")"]
        ]
        
        for row_offset, row_data in enumerate(chart_data):
            for col_offset, value in enumerate(row_data):
                cell = ws.cell(row=start_row + 1 + row_offset, column=1 + col_offset, value=value)
                if row_offset == 0:  # 헤더
                    cell.font = Font(name='Arial', size=10, bold=True)
                    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    def save_workbook(self, wb: Workbook, file_path: Union[str, Path]) -> str:
        """워크북을 파일로 저장"""
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb.save(file_path)
        logger.info(f"Excel workbook saved to: {file_path}")
        return str(file_path)
    
    def export_to_bytes(self, wb: Workbook) -> bytes:
        """워크북을 바이트로 내보내기 (Streamlit 다운로드용)"""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def create_editable_dataframe(self, scenarios: List[ExcelTestScenario]) -> pd.DataFrame:
        """Streamlit 편집용 DataFrame 생성"""
        if not scenarios:
            return self.template.create_empty_dataframe()
        
        data = [scenario.to_dict() for scenario in scenarios]
        return pd.DataFrame(data)
    
    def dataframe_to_scenarios(self, df: pd.DataFrame) -> List[ExcelTestScenario]:
        """DataFrame을 ExcelTestScenario 리스트로 변환"""
        scenarios = []
        for _, row in df.iterrows():
            try:
                scenario = ExcelTestScenario.from_dict(row.to_dict())
                scenarios.append(scenario)
            except Exception as e:
                logger.warning(f"Failed to convert row to scenario: {e}")
                continue
        return scenarios
    
    def generate_from_dataframe(
        self, 
        df: pd.DataFrame, 
        project_info: Optional[Dict[str, str]] = None
    ) -> Workbook:
        """DataFrame에서 직접 워크북 생성 (Streamlit용)"""
        scenarios = self.dataframe_to_scenarios(df)
        return self.generate_workbook(scenarios, project_info)
    
    def get_default_project_info(self) -> Dict[str, str]:
        """기본 프로젝트 정보"""
        return {
            "project_name": "AI Generated Test Project",
            "version": "1.0.0",
            "test_environment": "Development",
            "test_period": datetime.now().strftime("%Y-%m-%d"),
            "tester": "AI Test Generator"
        }