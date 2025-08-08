"""
Excel Templates Module - 엑셀 문서 템플릿 정의

테스트 시나리오 엑셀 문서의 구조와 스타일을 정의합니다.
Streamlit과 CLI 모두에서 사용 가능한 템플릿을 제공합니다.
"""
from dataclasses import dataclass, field
from typing import Dict, List, Any, Optional
from enum import Enum
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook


class TestPriority(str, Enum):
    """테스트 우선순위"""
    HIGH = "High"
    MEDIUM = "Medium"
    LOW = "Low"


class TestType(str, Enum):
    """테스트 타입"""
    FUNCTIONAL = "Functional"
    INTEGRATION = "Integration"
    PERFORMANCE = "Performance"
    SECURITY = "Security"
    REGRESSION = "Regression"


class TestStatus(str, Enum):
    """테스트 상태"""
    NOT_EXECUTED = "Not Executed"
    PASS = "Pass"
    FAIL = "Fail"
    BLOCKED = "Blocked"
    SKIP = "Skip"


@dataclass
class ExcelTestScenario:
    """Streamlit 편집 가능한 테스트 시나리오 데이터 모델"""
    scenario_id: str
    feature: str
    description: str
    preconditions: str  # 개행으로 구분된 문자열
    test_steps: str     # 개행으로 구분된 단계별 설명
    expected_results: str  # 개행으로 구분된 기대 결과
    test_data: str = ""    # JSON 문자열 또는 설명
    priority: str = TestPriority.MEDIUM.value
    test_type: str = TestType.FUNCTIONAL.value
    status: str = TestStatus.NOT_EXECUTED.value
    assigned_to: str = ""
    estimated_time: str = ""  # 예상 소요 시간 (분)
    actual_time: str = ""     # 실제 소요 시간 (분)
    notes: str = ""
    
    def to_dict(self) -> Dict[str, Any]:
        """Streamlit ag-grid용 딕셔너리 변환"""
        return {
            "Scenario ID": self.scenario_id,
            "Feature": self.feature,
            "Description": self.description,
            "Preconditions": self.preconditions,
            "Test Steps": self.test_steps,
            "Expected Results": self.expected_results,
            "Test Data": self.test_data,
            "Priority": self.priority,
            "Test Type": self.test_type,
            "Status": self.status,
            "Assigned To": self.assigned_to,
            "Estimated Time (min)": self.estimated_time,
            "Actual Time (min)": self.actual_time,
            "Notes": self.notes
        }
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'ExcelTestScenario':
        """딕셔너리에서 객체 생성"""
        return cls(
            scenario_id=str(data.get("Scenario ID", "")),
            feature=str(data.get("Feature", "")),
            description=str(data.get("Description", "")),
            preconditions=str(data.get("Preconditions", "")),
            test_steps=str(data.get("Test Steps", "")),
            expected_results=str(data.get("Expected Results", "")),
            test_data=str(data.get("Test Data", "")),
            priority=str(data.get("Priority", TestPriority.MEDIUM.value)),
            test_type=str(data.get("Test Type", TestType.FUNCTIONAL.value)),
            status=str(data.get("Status", TestStatus.NOT_EXECUTED.value)),
            assigned_to=str(data.get("Assigned To", "")),
            estimated_time=str(data.get("Estimated Time (min)", "")),
            actual_time=str(data.get("Actual Time (min)", "")),
            notes=str(data.get("Notes", ""))
        )
    
    @classmethod
    def from_test_scenario(cls, scenario, scenario_id: str = None) -> 'ExcelTestScenario':
        """기존 TestScenario 객체에서 변환"""
        return cls(
            scenario_id=scenario_id or scenario.scenario_id,
            feature=scenario.feature,
            description=scenario.description,
            preconditions="\n".join(scenario.preconditions) if scenario.preconditions else "",
            test_steps="\n".join([
                f"{i+1}. {step.get('action', '')}: {step.get('data', '')}"
                for i, step in enumerate(scenario.test_steps)
            ]) if scenario.test_steps else "",
            expected_results="\n".join(scenario.expected_results) if scenario.expected_results else "",
            test_data=str(scenario.test_data) if scenario.test_data else "",
            priority=scenario.priority,
            test_type=scenario.test_type
        )


@dataclass
class ExcelStyles:
    """엑셀 스타일 정의"""
    
    # 폰트
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    content_font = Font(name='Arial', size=10)
    
    # 배경색
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    priority_high_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    priority_medium_fill = PatternFill(start_color='FFE66D', end_color='FFE66D', fill_type='solid')
    priority_low_fill = PatternFill(start_color='95E1D3', end_color='95E1D3', fill_type='solid')
    
    status_pass_fill = PatternFill(start_color='A8E6CF', end_color='A8E6CF', fill_type='solid')
    status_fail_fill = PatternFill(start_color='FFB3BA', end_color='FFB3BA', fill_type='solid')
    status_blocked_fill = PatternFill(start_color='FFDFBA', end_color='FFDFBA', fill_type='solid')
    
    # 테두리
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 정렬
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)


class ExcelTemplate:
    """엑셀 템플릿 생성 및 관리"""
    
    @staticmethod
    def get_column_definitions() -> List[Dict[str, Any]]:
        """Streamlit ag-grid용 컬럼 정의"""
        return [
            {
                "headerName": "Scenario ID",
                "field": "Scenario ID",
                "width": 120,
                "pinned": "left",
                "editable": True
            },
            {
                "headerName": "Feature",
                "field": "Feature",
                "width": 150,
                "editable": True
            },
            {
                "headerName": "Description",
                "field": "Description",
                "width": 250,
                "editable": True,
                "cellEditor": "agLargeTextCellEditor"
            },
            {
                "headerName": "Preconditions",
                "field": "Preconditions",
                "width": 200,
                "editable": True,
                "cellEditor": "agLargeTextCellEditor"
            },
            {
                "headerName": "Test Steps",
                "field": "Test Steps",
                "width": 300,
                "editable": True,
                "cellEditor": "agLargeTextCellEditor"
            },
            {
                "headerName": "Expected Results",
                "field": "Expected Results",
                "width": 250,
                "editable": True,
                "cellEditor": "agLargeTextCellEditor"
            },
            {
                "headerName": "Test Data",
                "field": "Test Data",
                "width": 150,
                "editable": True
            },
            {
                "headerName": "Priority",
                "field": "Priority",
                "width": 100,
                "editable": True,
                "cellEditor": "agSelectCellEditor",
                "cellEditorParams": {
                    "values": [e.value for e in TestPriority]
                }
            },
            {
                "headerName": "Test Type",
                "field": "Test Type",
                "width": 120,
                "editable": True,
                "cellEditor": "agSelectCellEditor",
                "cellEditorParams": {
                    "values": [e.value for e in TestType]
                }
            },
            {
                "headerName": "Status",
                "field": "Status",
                "width": 120,
                "editable": True,
                "cellEditor": "agSelectCellEditor",
                "cellEditorParams": {
                    "values": [e.value for e in TestStatus]
                }
            },
            {
                "headerName": "Assigned To",
                "field": "Assigned To",
                "width": 120,
                "editable": True
            },
            {
                "headerName": "Estimated Time (min)",
                "field": "Estimated Time (min)",
                "width": 150,
                "editable": True,
                "type": "numericColumn"
            },
            {
                "headerName": "Actual Time (min)",
                "field": "Actual Time (min)",
                "width": 140,
                "editable": True,
                "type": "numericColumn"
            },
            {
                "headerName": "Notes",
                "field": "Notes",
                "width": 200,
                "editable": True,
                "cellEditor": "agLargeTextCellEditor"
            }
        ]
    
    @staticmethod
    def create_empty_dataframe() -> pd.DataFrame:
        """Streamlit 편집용 빈 DataFrame 생성"""
        columns = [col["field"] for col in ExcelTemplate.get_column_definitions()]
        
        # 샘플 데이터 1개 행 추가
        sample_data = {
            "Scenario ID": "TC001",
            "Feature": "User Authentication",
            "Description": "Test user login functionality",
            "Preconditions": "1. User account exists\n2. Application is running",
            "Test Steps": "1. Navigate to login page\n2. Enter valid credentials\n3. Click login button",
            "Expected Results": "1. User is logged in successfully\n2. Redirected to dashboard",
            "Test Data": "username: test@example.com\npassword: Test123!",
            "Priority": TestPriority.HIGH.value,
            "Test Type": TestType.FUNCTIONAL.value,
            "Status": TestStatus.NOT_EXECUTED.value,
            "Assigned To": "",
            "Estimated Time (min)": "5",
            "Actual Time (min)": "",
            "Notes": ""
        }
        
        df = pd.DataFrame([sample_data])
        return df
    
    @staticmethod
    def get_summary_template() -> Dict[str, Any]:
        """테스트 요약 시트 템플릿"""
        return {
            "title": "Test Execution Summary",
            "sections": [
                {
                    "name": "Project Information",
                    "fields": [
                        ("Project Name", ""),
                        ("Version", ""),
                        ("Test Environment", ""),
                        ("Test Period", ""),
                        ("Tester", "")
                    ]
                },
                {
                    "name": "Test Statistics",
                    "fields": [
                        ("Total Test Cases", "=COUNTA(TestScenarios!A:A)-1"),
                        ("Executed", "=COUNTIF(TestScenarios!J:J,\"Pass\")+COUNTIF(TestScenarios!J:J,\"Fail\")"),
                        ("Passed", "=COUNTIF(TestScenarios!J:J,\"Pass\")"),
                        ("Failed", "=COUNTIF(TestScenarios!J:J,\"Fail\")"),
                        ("Blocked", "=COUNTIF(TestScenarios!J:J,\"Blocked\")"),
                        ("Not Executed", "=COUNTIF(TestScenarios!J:J,\"Not Executed\")")
                    ]
                },
                {
                    "name": "Priority Breakdown",
                    "fields": [
                        ("High Priority", "=COUNTIF(TestScenarios!H:H,\"High\")"),
                        ("Medium Priority", "=COUNTIF(TestScenarios!H:H,\"Medium\")"),
                        ("Low Priority", "=COUNTIF(TestScenarios!H:H,\"Low\")")
                    ]
                }
            ]
        }